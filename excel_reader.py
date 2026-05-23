"""Excel file reading and data extraction module."""
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from config import (
    EXCEL_FILE,
    SHEET_PROJECT,
    SHEET_SUMMARY,
    SHEET_RESPONSE,
    SHEET_TEAM,
    SUMMARY_HEADER_ROW,
    SUMMARY_REQUEST_ID_COL,
    SUMMARY_CONTENT_COL,
    SUMMARY_DEADLINE_COL,
    SUMMARY_STATUS_COL,
    RESPONSE_HEADER_ROW,
    RESPONSE_REQUEST_ID_COL,
    RESPONSE_TEAM_NO_COL,
    RESPONSE_CONTENT_COL,
    TEAM_HEADER_ROW,
    TEAM_NO_COL,
    TEAM_CONTACT_START_COL,
    TEAM_CONTACT_END_COL,
    TARGET_STATUS,
)
from utils import setup_logger, clean_value, get_reminder_timing
from datetime import datetime

logger = setup_logger(__name__)


class ExcelReader:
    """Read and extract data from Excel file."""

    def __init__(self):
        """Initialize ExcelReader."""
        try:
            self.workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
            self.sheet_project = self.workbook.worksheets[SHEET_PROJECT]
            self.sheet_summary = self.workbook.worksheets[SHEET_SUMMARY]
            self.sheet_response = self.workbook.worksheets[SHEET_RESPONSE]
            self.sheet_team = self.workbook.worksheets[SHEET_TEAM]
            logger.info(f"Loaded Excel file: {EXCEL_FILE}")
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise

    def get_pending_requests(self) -> list[dict]:
        """Extract request information with status='依頼中' from summary sheet."""
        requests = []
        try:
            for row in range(SUMMARY_HEADER_ROW + 1, self.sheet_summary.max_row + 1):
                status = clean_value(
                    self.sheet_summary.cell(row, SUMMARY_STATUS_COL).value
                )
                if status == TARGET_STATUS:
                    request_id = clean_value(
                        self.sheet_summary.cell(row, SUMMARY_REQUEST_ID_COL).value
                    )
                    if request_id:
                        content = clean_value(
                            self.sheet_summary.cell(row, SUMMARY_CONTENT_COL).value
                        )
                        deadline = clean_value(
                            self.sheet_summary.cell(row, SUMMARY_DEADLINE_COL).value
                        )
                        requests.append({
                            "id": request_id,
                            "content": content,
                            "deadline": deadline,
                        })
            logger.info(f"Found {len(requests)} pending requests")
            return requests
        except Exception as e:
            logger.error(f"Error extracting pending requests: {e}")
            raise

    def get_pending_requests_by_deadline_timing(self) -> dict:
        """
        Extract pending requests filtered by deadline timing.

        Returns:
            dict with keys '3days_before', '1day_before', 'on_deadline'
            each containing list of request dicts with 'id', 'content', 'deadline'
        """
        all_requests = self.get_pending_requests()
        timing_dict = {
            "3days_before": [],
            "1day_before": [],
            "on_deadline": [],
        }

        today = datetime.now()

        try:
            for req in all_requests:
                deadline_str = req["deadline"]
                if not deadline_str:
                    continue

                deadline_date = None
                # Try to parse as datetime object (if already parsed by openpyxl)
                if isinstance(deadline_str, datetime):
                    deadline_date = deadline_str
                else:
                    # Try various string formats
                    formats = [
                        "%Y-%m-%d %H:%M:%S",
                        "%Y/%m/%d %H:%M:%S",
                        "%Y-%m-%d",
                        "%Y/%m/%d",
                    ]
                    for fmt in formats:
                        try:
                            deadline_date = datetime.strptime(deadline_str, fmt)
                            break
                        except (ValueError, TypeError):
                            continue

                    if not deadline_date:
                        logger.warning(
                            f"Could not parse deadline for request {req['id']}: {deadline_str}"
                        )
                        continue

                if deadline_date:
                    timing = get_reminder_timing(deadline_date, today)
                    if timing:
                        timing_dict[timing].append(req)
                        logger.info(
                            f"Request {req['id']} matched timing: {timing}"
                        )

            return timing_dict
        except Exception as e:
            logger.error(f"Error filtering requests by deadline timing: {e}")
            raise

    def get_teams_needing_response(self, requests: list[dict]) -> list[dict]:
        """
        Extract teams needing response with their associated requests.

        Args:
            requests: List of request dicts with 'id', 'content', 'deadline'

        Returns:
            List of dicts with 'team_no' and 'requests'
        """
        request_ids = [req["id"] for req in requests]
        request_map = {req["id"]: req for req in requests}
        team_requests = {}

        try:
            for row in range(RESPONSE_HEADER_ROW + 1, self.sheet_response.max_row + 1):
                req_id = clean_value(
                    self.sheet_response.cell(row, RESPONSE_REQUEST_ID_COL).value
                )
                response_content = clean_value(
                    self.sheet_response.cell(row, RESPONSE_CONTENT_COL).value
                )

                if req_id in request_ids and response_content == "":
                    team_no = clean_value(
                        self.sheet_response.cell(row, RESPONSE_TEAM_NO_COL).value
                    )
                    if team_no:
                        if team_no not in team_requests:
                            team_requests[team_no] = []
                        team_requests[team_no].append(request_map[req_id])
                        logger.info(
                            f"Found team needing response: {team_no} (request ID: {req_id})"
                        )

            result = [
                {"team_no": team_no, "requests": reqs}
                for team_no, reqs in team_requests.items()
            ]
            logger.info(f"Total unique teams needing response: {len(result)}")
            return result
        except Exception as e:
            logger.error(f"Error extracting teams needing response: {e}")
            raise

    def get_team_contacts(self, team_no: str) -> list[str]:
        """Extract email addresses for a given team number."""
        contacts = []
        try:
            for row in range(TEAM_HEADER_ROW + 1, self.sheet_team.max_row + 1):
                current_team_no = clean_value(
                    self.sheet_team.cell(row, TEAM_NO_COL).value
                )
                if current_team_no == team_no:
                    # Extract all contact columns (C to G)
                    for col in range(TEAM_CONTACT_START_COL, TEAM_CONTACT_END_COL + 1):
                        email = clean_value(self.sheet_team.cell(row, col).value)
                        if email and email != "":
                            contacts.append(email)
                    logger.info(
                        f"Found {len(contacts)} contacts for team {team_no}: {contacts}"
                    )
                    break
            else:
                logger.warning(f"Team {team_no} not found in team definition sheet")

            return contacts
        except Exception as e:
            logger.error(f"Error extracting team contacts for {team_no}: {e}")
            raise

    def get_project_info(self) -> dict:
        """Extract project information from PJ情報 sheet."""
        try:
            from config import (
                PROJECT_NAME_ROW,
                PROJECT_PM_EMAIL_ROW,
                PROJECT_CC_EMAIL_ROW,
                PROJECT_VALUE_COL,
            )

            project_name = clean_value(
                self.sheet_project.cell(PROJECT_NAME_ROW, PROJECT_VALUE_COL).value
            )
            pm_email = clean_value(
                self.sheet_project.cell(PROJECT_PM_EMAIL_ROW, PROJECT_VALUE_COL).value
            )
            cc_email = clean_value(
                self.sheet_project.cell(PROJECT_CC_EMAIL_ROW, PROJECT_VALUE_COL).value
            )

            result = {
                "project_name": project_name,
                "pm_email": pm_email,
                "cc_email": cc_email,
            }
            logger.info(f"Project info: {result}")
            return result
        except Exception as e:
            logger.error(f"Error extracting project info: {e}")
            raise

    def close(self):
        """Close workbook."""
        self.workbook.close()
        logger.info("Excel file closed")
